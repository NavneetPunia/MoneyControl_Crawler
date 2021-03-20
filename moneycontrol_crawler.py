#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import xlsxwriter
import pandas as pd
from bs4 import BeautifulSoup
from urllib.request import urlopen
from urllib.request import Request, URLError
from openpyxl import load_workbook
import time
import random
from datetime import datetime



header = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:32.0) Gecko/20100101 Firefox/32.0',}


def read_page_no(file_name):
    file = open(file_name, "r")
    page_no = file.readline()
    file.close()
    return int(page_no)


def write_page_no(file_name, page_no):
    file = open(file_name, "w+")
    file.write(page_no)
    file.close()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, index=False, header=False)

    # save the workbook
    writer.save()


def delay() -> None:
    time.sleep(random.uniform(15, 30))
    return None

def get_districts(url):
    req = Request(url, headers=header)
    #print(req)
    webpage = urlopen(req).read()
    parser = BeautifulSoup(webpage, features="html.parser")
    df = pd.DataFrame(columns=['date_time', 'Heading'])

    ul_list = parser.find('ul', attrs={'id': 'cagetory'})
    for rows in ul_list.find_all('li', attrs={'class': 'clearfix'}):
        heading_text = rows.find('h2').text
        date_text = rows.find('span').text
        df.loc[df.shape[0]] = [str(date_text), str(heading_text)]
        print("DATE AND TIME: ", date_text, " ,HEADING: ", heading_text)

    print("\n")
    append_df_to_excel("heading/MONEY_CONTROL_DATASET.xlsx", df)
    #print(ul_list)

#Total pages number are 359-990, to make dataset mentioned in Synopsis

if __name__ == '__main__':
    page_no = read_page_no('page_no.txt')
    counter = 1
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    print("Start Time =", current_time)

    while(int(page_no)<=990):
        if counter >= 101: # Counter to Crawl only 100 Moneycontrol Pages, if exceed break the crawling
            break
        else:
            page_no = read_page_no('page_no.txt')
            url = 'https://www.moneycontrol.com/news/business/economy/page-' + str(page_no)
            print(url)
            get_districts(url)
            write_page_no('page_no.txt', str(page_no + 1))
            counter = counter+1
            delay()
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    print("End Time =", current_time)


    '''
    page_no = read_page_no('page_no.txt')
    print(page_no)
    write_page_no('page_no.txt', str(page_no+1))
    page_no = read_page_no('page_no.txt')
    print(page_no)
    '''
